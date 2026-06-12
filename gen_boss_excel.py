"""生成孤云支线18Boss数值Excel"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "孤云支线18Boss"

# === 样式 ===
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
boss_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
normal_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

# === 数据 ===
bosses = [
    # id, name, icon, hp, qi, atk, def, combo, hit, dodge, crit, speed, rank, boss, trait, traitDesc, taunt
    ("M2", "堂口捕头·刘铁", "捕", 500, 180, 40, 16, 3, 65, 5, 6, 1.15, 1, False, "", "", "堂口有令——拒册散人，押回问话！"),
    ("M4", "缉捕队长·钱虎", "缉", 750, 240, 55, 22, 3, 68, 5, 7, 1.20, 1, False, "", "", "老子在堂口十年，还没哪个散人跑得掉。"),
    ("M6", "铁手·周通", "拳", 1000, 320, 70, 28, 5, 70, 6, 10, 1.25, 2, True, "armorBreak", "铁手套开瓢，拳拳破防", "奉命'劝导'拒册散人。劝不听的话——我这双手套开过不少瓢。"),
    ("M8", "先锋营统领·马如龙", "将", 1500, 420, 85, 34, 4, 72, 7, 9, 1.30, 2, False, "", "", "先锋营在此！散人还不束手就擒？"),
    ("M10", "护法副将·杨震", "将", 2000, 520, 100, 40, 4, 75, 8, 10, 1.35, 3, True, "miniArmor", "护体真气，高防稳守", "左护法点名要你的人头。自己交出来，免得多受皮肉苦。"),
    ("M12", "杭州堂主·赵崇岳", "刀", 4000, 1200, 120, 48, 5, 80, 10, 14, 1.50, 5, True, "lowHpBerserk", "低血时攻速双升；九环刀法范围攻击", "知不知道因为你一个人，我少赚了多少银子？"),
    ("M14", "沈千山帐前哨长·杜威", "哨", 2500, 600, 110, 44, 4, 74, 9, 10, 1.40, 3, False, "", "", "左护法的眼睛无处不在。你藏不住的。"),
    ("M16", "寒剑·柳长卿", "剑", 3000, 750, 125, 50, 5, 76, 14, 14, 1.50, 3, True, "miniFrost", "寒霜剑气，减速削内", "你的剑法，比传闻中弱。"),
    ("M18", "夜袭队长·秦烈", "袭", 3500, 780, 140, 56, 5, 78, 10, 12, 1.48, 4, False, "", "", "夜长梦多——速战速决，一个不留。"),
    ("M20", "「血手」崔命", "血", 4000, 850, 155, 62, 4, 80, 12, 14, 1.52, 4, True, "miniBleed", "链子锤重创，流血+2", "五千两是你的命价——但我不急着收，先玩玩。"),
    ("M22", "无影·叶孤", "影", 4500, 1000, 170, 68, 7, 88, 30, 18, 1.80, 4, True, "highDodge", "极速暗杀，闪避+15", "我要的不是你的命，是那份名单。交出来，你可以活。"),
    ("M24", "左护法·沈千山", "戟", 8000, 2400, 200, 80, 6, 88, 14, 18, 1.65, 7, True, "berserkSummon", "70%血狂暴；30%血召唤护卫", "把所有人当资源配置——包括你我。区别只是价格不同。"),
    ("M26", "狂刀·钱彪", "刀", 5000, 1100, 185, 74, 5, 82, 10, 14, 1.60, 5, True, "lowHpBerserk", "低血狂暴，攻速双升", "统领建立的新秩序，需要你们这些散人做出点牺牲。"),
    ("M28", "精英卫队长·卫岳", "卫", 5500, 1200, 200, 80, 5, 84, 10, 15, 1.65, 5, False, "", "", "总坛禁卫在此。擅入者，踏过我的尸体。"),
    ("M30", "烽火统领·霍烽", "烽", 6000, 1400, 215, 86, 5, 85, 9, 16, 1.70, 6, True, "armorBreak", "狼牙棒重击，破防贯通", "太行外围百里之内，没有我的狼烟传不到的信号。"),
    ("M32", "右护法·公孙烈", "枪", 6500, 1600, 230, 92, 6, 85, 10, 16, 1.75, 6, True, "armorBreak", "浑铁枪破阵，防御贯通", "统领说打谁就打谁。我不问为什么。"),
    ("M34", "地牢典狱长·阎铁", "狱", 7000, 1800, 245, 98, 4, 88, 12, 14, 1.80, 6, True, "pointStrike", "判官笔打穴，概率封行动", "来了就别走了。地牢的铁链还有空位。"),
    ("M36", "武盟统领·楚宗玄", "魔", 15000, 4000, 280, 112, 8, 95, 20, 24, 2.00, 10, True, "shieldPurityBerserk", "开场25%护体；50%血净化；15%血攻翻倍防归零", "维持一个能救千万人的机构需要代价。每年几百个散人变成数字——我觉得值得。"),
]

# === 列定义 ===
headers = ["月份", "名称", "图标", "rank", "HP", "QI", "ATK", "DEF", "COMBO", "HIT", "DODGE", "CRIT", "SPEED", "Boss特性代码", "特性描述"]
col_widths = [6, 24, 5, 6, 8, 8, 8, 8, 8, 8, 8, 8, 8, 24, 36]

# === 写入 ===
# 标题行
ws.merge_cells("A1:O1")
title_cell = ws["A1"]
title_cell.value = "孤云逐浪 · 孤云支线18Boss九维与特性总表"
title_cell.font = Font(name="微软雅黑", size=16, bold=True, color="1F3864")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

# 表头
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
ws.row_dimensions[2].height = 24

# 数据行
for row_idx, b in enumerate(bosses, 3):
    is_boss = b[11]
    values = [
        b[0],  # 月份
        b[1],  # 名称
        b[2],  # 图标
        b[11],  # rank
        b[3],   # HP
        b[4],   # QI
        b[5],   # ATK
        b[6],   # DEF
        b[7],   # COMBO
        b[8],   # HIT
        b[9],   # DODGE
        b[10],  # CRIT
        b[12],  # SPEED
        b[13] if is_boss else "—",   # trait代码
        b[14] if is_boss else "—",   # 中文描述
    ]
    fill = boss_fill if is_boss else normal_fill
    for col_idx, v in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=v)
        cell.font = Font(name="微软雅黑", size=10)
        cell.fill = fill
        cell.border = thin_border
        if col_idx in (1, 3, 4):
            cell.alignment = center_align
        elif col_idx in (2, 14, 15):
            cell.alignment = left_align
        else:
            cell.alignment = center_align
    ws.row_dimensions[row_idx].height = 22

# 列宽
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# === 特性代码对照 Sheet ===
ws2 = wb.create_sheet("特性代码对照")

trait_data = [
    ("armorBreak", "拳拳破防，无视敌方一定防御"),
    ("miniArmor", "护体真气，额外护甲减伤"),
    ("lowHpBerserk", "低血量时攻击力与速度显著提升"),
    ("miniFrost", "寒霜剑气，减速并削减内力"),
    ("miniBleed", "链子锤重创，附加流血debuff"),
    ("highDodge", "极速身法，额外闪避加成"),
    ("berserkSummon", "多阶段狂暴+召唤护卫"),
    ("pointStrike", "判官笔打穴，概率封印行动"),
    ("shieldPurityBerserk", "护体→净化→极限狂暴三段式"),
]
ws2.merge_cells("A1:C1")
ws2["A1"].value = "Boss特性代码对照表"
ws2["A1"].font = Font(name="微软雅黑", size=14, bold=True, color="1F3864")
ws2["A1"].alignment = Alignment(horizontal="center")
ws2.row_dimensions[1].height = 30

for ci, h in enumerate(["特性代码", "出现Boss", "效果描述"], 1):
    cell = ws2.cell(row=2, column=ci, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    cell.alignment = center_align
    cell.border = thin_border

# 统计各trait出现次数
from collections import Counter
trait_count = Counter()
for b in bosses:
    if b[13]:
        trait_count[b[13]] += 1

trait_bosses = {t: [b[0] for b in bosses if b[13] == t] for t in trait_count}

for ri, (code, desc) in enumerate(trait_data, 3):
    bosses_str = "、".join(trait_bosses.get(code, []))
    for ci, v in enumerate([code, bosses_str, desc], 1):
        cell = ws2.cell(row=ri, column=ci, value=v)
        cell.font = Font(name="微软雅黑", size=10)
        cell.border = thin_border
        cell.alignment = left_align if ci > 1 else center_align
    ws2.row_dimensions[ri].height = 22

ws2.column_dimensions["A"].width = 24
ws2.column_dimensions["B"].width = 28
ws2.column_dimensions["C"].width = 40

# === 九维说明 Sheet ===
ws3 = wb.create_sheet("九维说明")
ws3.merge_cells("A1:B1")
ws3["A1"].value = "九维属性说明"
ws3["A1"].font = Font(name="微软雅黑", size=14, bold=True, color="1F3864")
ws3["A1"].alignment = Alignment(horizontal="center")
ws3.row_dimensions[1].height = 28

stats_info = [
    ("HP", "生命值 — 归零即战败"),
    ("QI", "内力值 — 释放技能消耗"),
    ("ATK", "攻击力 — 基础伤害"),
    ("DEF", "防御力 — 减免伤害"),
    ("COMBO", "连击数 — 每回合攻击次数上限"),
    ("HIT", "命中 — 决定攻击是否命中"),
    ("DODGE", "闪避 — 决定能否闪避攻击"),
    ("CRIT", "暴击率 — 触发暴击的概率"),
    ("SPEED", "速度 — 决定出手顺序"),
    ("rank", "位阶 — 1~10，影响掉落和难度"),
]
for ri, (k, v) in enumerate(stats_info, 2):
    ws3.cell(row=ri, column=1, value=k).font = Font(name="微软雅黑", size=10, bold=True)
    ws3.cell(row=ri, column=2, value=v).font = Font(name="微软雅黑", size=10)
    ws3.cell(row=ri, column=1).alignment = center_align
    ws3.cell(row=ri, column=2).alignment = left_align
    for ci in (1, 2):
        ws3.cell(row=ri, column=ci).border = thin_border

ws3.column_dimensions["A"].width = 12
ws3.column_dimensions["B"].width = 36

# === 冻结窗格 ===
ws.freeze_panes = "A3"

# 保存
out_path = "C:/Users/Administrator.DESKTOP-O2PR2VT/Desktop/wuxiasimulator/孤云支线18Boss数值表.xlsx"
wb.save(out_path)
print(f"OK: {out_path}")
